from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
import time

def get_data_from_excel(file_path, sheet_name, row_number):
    data = {}
    
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Select the desired sheet
    sheet = workbook[sheet_name]
    while True:
        status_cell = sheet.cell(row=row_number, column=sheet['F'][0].column)  # Assuming the status is in column F
        if status_cell.value != "Done":
            # If status is not "done," break the loop
            break
        row_number += 1
    
    if status_cell.value != "Done":
        # Get column names from the first row
        column_names = [cell.value for cell in sheet[1]]
        
        # Get values from the specified row
        row_values = [cell.value for cell in sheet[row_number]]
        
        # Create a dictionary with column names as keys and corresponding values
        data = dict(zip(column_names, row_values))
    
    print(f"Row Number: {row_number}")
    # Close the workbook
    workbook.close()
    
    return data, row_number  # Return both data and row_number

def fill_form_with_excel_data(row_number):
    driver = None
    try:
        # Open Chrome browser
        driver = webdriver.Chrome()
        driver.get('http://localhost:5000/')
        #to make sure it stays for a while
        time.sleep(2)
        #path to the excel file
        excel_file_path = r"D:\Self_learning\registration.xlsx"
        sheet_name = "details"

        employee_data, row_number = get_data_from_excel(excel_file_path, sheet_name, row_number)

        if employee_data:
            # mapping between form field names and Excel column names
            field_mapping = {
                "userName": "Name",
                "address": "Address",
                "contact": "ContactNumber",
                "department": "Department"
            }

            # Fill in form fields
            for form_field, excel_column in field_mapping.items():
                if excel_column in employee_data:
                    input_field = driver.find_element("name", form_field)
                    input_field.send_keys(employee_data[excel_column])

            # Submit the form
            submit = driver.find_element("name", "register")
            submit.click()

            # Update status in the Excel sheet
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook[sheet_name]
            status_cell = sheet.cell(row=row_number, column=sheet['F'][0].column) 
            status_cell.value = "Done"
            workbook.save(excel_file_path)

    except Exception as e:
        print(f"An exception occurred: {str(e)}")

    finally:
        # Close the WebDriver instance
        if driver:
            driver.quit()

if __name__ == "__main__":
    row_number = 2  # Set an initial row number
    fill_form_with_excel_data(row_number)
